from __future__ import annotations

import asyncio
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

from nexus.config import Settings
from nexus.tools.excel import ExcelTool


class _FakeRecalcEngine:
    def __init__(self, report: dict | None = None) -> None:
        self.report = report or {
            "ok": True,
            "status": "success",
            "formula_count": 1,
            "total_errors": 0,
            "error_summary": {},
            "recalc_applied": True,
        }
        self.calls: list[str] = []

    def recalc_and_validate(self, workbook_path: Path) -> dict:
        self.calls.append(str(workbook_path))
        return self.report


def _settings(tmp_path: Path, **kwargs) -> Settings:
    base = {
        "db_path": tmp_path / "nexus.db",
        "workspace": tmp_path / "workspace",
        "memories_dir": tmp_path / "memories",
    }
    base.update(kwargs)
    return Settings(**base)


def test_excel_create_requires_confirmation(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path))
    result = asyncio.run(tool.run({"action": "create", "path": "book.xlsx"}))
    assert not result.ok
    assert result.requires_confirmation


def test_excel_create_and_write_read(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())

    created = asyncio.run(
        tool.run({"action": "create", "path": "book.xlsx", "sheet_name": "Data", "confirmed": True})
    )
    assert created.ok
    assert created.artifacts

    write = asyncio.run(
        tool.run(
            {
                "action": "write_cells",
                "path": "book.xlsx",
                "sheet": "Data",
                "cells": {"A1": "Name", "B1": "Score", "A2": "Liam", "B2": 95},
                "confirmed": True,
            }
        )
    )
    assert write.ok

    read = asyncio.run(tool.run({"action": "read", "path": "book.xlsx", "sheet": "Data", "range": "A1:B2"}))
    assert read.ok
    assert "Liam" in read.content


def test_excel_append_rows(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    asyncio.run(tool.run({"action": "create", "path": "book.xlsx", "confirmed": True}))
    appended = asyncio.run(
        tool.run(
            {
                "action": "append_rows",
                "path": "book.xlsx",
                "rows": [["A", 1], ["B", 2]],
                "confirmed": True,
            }
        )
    )
    assert appended.ok


def test_excel_read_serializes_temporal_cells_to_iso(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    asyncio.run(tool.run({"action": "create", "path": "temporal.xlsx", "confirmed": True}))

    workbook_path = tmp_path / "workspace" / "temporal.xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws["A1"] = datetime(2026, 2, 18, 22, 13, 10)
    ws["B1"] = date(2026, 2, 18)
    ws["C1"] = time(9, 45, 0)
    ws["D1"] = timedelta(days=1, hours=2, minutes=3)
    workbook.save(workbook_path)
    workbook.close()

    read = asyncio.run(tool.run({"action": "read", "path": "temporal.xlsx", "range": "A1:D1"}))
    assert read.ok
    assert "2026-02-18T22:13:10" in read.content
    assert "2026-02-18" in read.content
    assert "09:45:00" in read.content
    assert "1 day, 2:03:00" in read.content


def test_excel_convert_xlsx_to_csv_and_back(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    asyncio.run(tool.run({"action": "create", "path": "book.xlsx", "confirmed": True}))
    asyncio.run(
        tool.run(
            {
                "action": "write_cells",
                "path": "book.xlsx",
                "cells": {"A1": "name", "B1": "score", "A2": "liam", "B2": 99},
                "confirmed": True,
            }
        )
    )

    to_csv = asyncio.run(
        tool.run(
            {
                "action": "convert",
                "path": "book.xlsx",
                "output_path": "book.csv",
                "confirmed": True,
            }
        )
    )
    assert to_csv.ok

    to_xlsx = asyncio.run(
        tool.run(
            {
                "action": "convert",
                "path": "book.csv",
                "output_path": "book-roundtrip.xlsx",
                "confirmed": True,
            }
        )
    )
    assert to_xlsx.ok


def test_excel_clean_table_normalizes_headers(tmp_path: Path):
    csv_path = tmp_path / "workspace" / "messy.csv"
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    csv_path.write_text("Name , Revenue ($mm),\n Liam , 100\n,\n", encoding="utf-8")

    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    cleaned = asyncio.run(
        tool.run(
            {
                "action": "clean_table",
                "path": "messy.csv",
                "output_path": "cleaned.xlsx",
                "normalize_headers": True,
                "drop_empty_rows": True,
                "drop_empty_cols": True,
                "confirmed": True,
            }
        )
    )
    assert cleaned.ok

    wb = load_workbook(tmp_path / "workspace" / "cleaned.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    wb.close()
    assert headers[0] == "name"
    assert headers[1] == "revenue_mm"


def test_excel_number_format_style_comment_and_chart(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    asyncio.run(tool.run({"action": "create", "path": "styled.xlsx", "confirmed": True}))
    asyncio.run(
        tool.run(
            {
                "action": "write_cells",
                "path": "styled.xlsx",
                "cells": {
                    "A1": "Month",
                    "B1": "Revenue",
                    "A2": "Jan",
                    "B2": 10,
                    "A3": "Feb",
                    "B3": 20,
                },
                "confirmed": True,
            }
        )
    )

    number_format = asyncio.run(
        tool.run(
            {
                "action": "set_number_format",
                "path": "styled.xlsx",
                "range": "B2:B3",
                "number_format": "$#,##0;($#,##0);-",
                "confirmed": True,
            }
        )
    )
    assert number_format.ok

    styled = asyncio.run(
        tool.run(
            {
                "action": "set_style",
                "path": "styled.xlsx",
                "range": "A1:B1",
                "preset": "professional",
                "bold": True,
                "confirmed": True,
            }
        )
    )
    assert styled.ok

    commented = asyncio.run(
        tool.run(
            {
                "action": "add_comment",
                "path": "styled.xlsx",
                "cell": "B2",
                "comment": "Source: Internal 2026-02-18",
                "confirmed": True,
            }
        )
    )
    assert commented.ok

    charted = asyncio.run(
        tool.run(
            {
                "action": "create_chart",
                "path": "styled.xlsx",
                "chart_type": "line",
                "data_range": "B1:B3",
                "category_range": "A2:A3",
                "position": "D2",
                "confirmed": True,
            }
        )
    )
    assert charted.ok

    wb = load_workbook(tmp_path / "workspace" / "styled.xlsx")
    ws = wb.active
    assert ws["B2"].number_format == "$#,##0;($#,##0);-"
    assert ws["A1"].font.bold is True
    assert ws["B2"].comment is not None
    assert len(ws._charts) == 1
    wb.close()


def test_excel_formula_write_triggers_recalc_and_strict_error(tmp_path: Path):
    recalc = _FakeRecalcEngine(
        {
            "ok": True,
            "status": "errors_found",
            "formula_count": 1,
            "total_errors": 1,
            "error_summary": {"#DIV/0!": {"count": 1, "locations": ["Sheet1!A1"]}},
            "recalc_applied": True,
        }
    )
    tool = ExcelTool(_settings(tmp_path, excel_strict_formula_errors=True), recalc_engine=recalc)
    asyncio.run(tool.run({"action": "create", "path": "formula.xlsx", "confirmed": True}))

    result = asyncio.run(
        tool.run(
            {
                "action": "write_cells",
                "path": "formula.xlsx",
                "cells": {"A1": "=1/0"},
                "confirmed": True,
            }
        )
    )
    assert not result.ok
    assert recalc.calls
    assert "Formula validation found Excel errors" in result.content


def test_excel_recalc_validate_action(tmp_path: Path):
    report = {
        "ok": True,
        "status": "success",
        "formula_count": 2,
        "total_errors": 0,
        "error_summary": {},
        "recalc_applied": True,
    }
    recalc = _FakeRecalcEngine(report)
    tool = ExcelTool(_settings(tmp_path), recalc_engine=recalc)
    asyncio.run(tool.run({"action": "create", "path": "validate.xlsx", "confirmed": True}))

    result = asyncio.run(
        tool.run(
            {
                "action": "recalc_validate",
                "path": "validate.xlsx",
                "confirmed": True,
            }
        )
    )
    assert result.ok
    assert "recalc_status=success" in result.content


def test_excel_xlsm_extension_preserved(tmp_path: Path):
    tool = ExcelTool(_settings(tmp_path), recalc_engine=_FakeRecalcEngine())
    created = asyncio.run(tool.run({"action": "create", "path": "macro.xlsm", "confirmed": True}))
    assert created.ok
    assert created.artifacts
    artifact = created.artifacts[0]
    assert artifact["mime_type"] == "application/vnd.ms-excel.sheet.macroEnabled.12"

    added = asyncio.run(
        tool.run(
            {
                "action": "add_sheet",
                "path": "macro.xlsm",
                "sheet_name": "Data",
                "confirmed": True,
            }
        )
    )
    assert added.ok
