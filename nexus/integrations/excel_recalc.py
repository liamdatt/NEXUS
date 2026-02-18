from __future__ import annotations

import shutil
import subprocess
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


EXCEL_ERROR_TOKENS = (
    "#VALUE!",
    "#DIV/0!",
    "#REF!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#N/A",
)


class ExcelRecalcEngine:
    def __init__(self, *, timeout_seconds: int = 45) -> None:
        self.timeout_seconds = max(1, int(timeout_seconds))

    def recalc_and_validate(self, workbook_path: Path) -> dict[str, Any]:
        path = workbook_path.expanduser().resolve()
        if not path.exists() or not path.is_file():
            return {"ok": False, "status": "error", "error": f"workbook not found: {path}"}

        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            return {"ok": False, "status": "error", "error": f"recalc not supported for extension: {suffix}"}

        formula_count = self._count_formulas(path)
        recalc_applied = False
        if formula_count > 0:
            recalc_applied = self._recalc_with_soffice(path)
            if not recalc_applied:
                return {
                    "ok": False,
                    "status": "error",
                    "error": "libreoffice recalculation failed; ensure soffice is available in runtime",
                    "formula_count": formula_count,
                }

        total_errors, error_summary = self._scan_formula_errors(path)
        status = "success" if total_errors == 0 else "errors_found"
        return {
            "ok": True,
            "status": status,
            "formula_count": formula_count,
            "total_errors": total_errors,
            "error_summary": error_summary,
            "recalc_applied": recalc_applied,
        }

    def _count_formulas(self, workbook_path: Path) -> int:
        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(workbook_path, data_only=False, keep_vba=keep_vba)
        try:
            count = 0
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if isinstance(value, str) and value.startswith("="):
                            count += 1
            return count
        finally:
            workbook.close()

    def _scan_formula_errors(self, workbook_path: Path) -> tuple[int, dict[str, dict[str, Any]]]:
        keep_vba = workbook_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(workbook_path, data_only=True, keep_vba=keep_vba)
        try:
            summary: dict[str, dict[str, Any]] = {}
            total = 0
            for token in EXCEL_ERROR_TOKENS:
                summary[token] = {"count": 0, "locations": []}

            for sheet in workbook.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        value = cell.value
                        if not isinstance(value, str):
                            continue
                        for token in EXCEL_ERROR_TOKENS:
                            if token not in value:
                                continue
                            total += 1
                            data = summary[token]
                            data["count"] = int(data["count"]) + 1
                            locations = data["locations"]
                            if isinstance(locations, list) and len(locations) < 50:
                                locations.append(f"{sheet.title}!{cell.coordinate}")
                            break

            compact = {k: v for k, v in summary.items() if int(v.get("count", 0)) > 0}
            return total, compact
        finally:
            workbook.close()

    def _recalc_with_soffice(self, workbook_path: Path) -> bool:
        suffix = workbook_path.suffix.lower()
        format_name = "xlsx"
        format_filter = "Calc MS Excel 2007 XML"
        if suffix == ".xlsm":
            format_name = "xlsm"
            format_filter = "Calc MS Excel 2007 VBA XML"

        with tempfile.TemporaryDirectory() as tmpdir:
            out_dir = Path(tmpdir)
            cmd = [
                "soffice",
                "--headless",
                "--norestore",
                "--nolockcheck",
                "--nodefault",
                "--convert-to",
                f"{format_name}:{format_filter}",
                "--outdir",
                str(out_dir),
                str(workbook_path),
            ]
            try:
                proc = subprocess.run(
                    cmd,
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=self.timeout_seconds,
                )
            except FileNotFoundError:
                return False
            except subprocess.TimeoutExpired:
                return False

            if proc.returncode != 0:
                return False

            direct = out_dir / workbook_path.name
            if direct.exists():
                shutil.copy2(direct, workbook_path)
                return True

            candidates = sorted(out_dir.glob(f"{workbook_path.stem}.*"))
            if not candidates:
                return False

            shutil.copy2(candidates[0], workbook_path)
            return True
