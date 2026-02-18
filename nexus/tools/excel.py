from __future__ import annotations

import json
import re
from copy import copy
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

from nexus.config import Settings
from nexus.integrations.excel_recalc import ExcelRecalcEngine
from nexus.tools.base import BaseTool, ToolResult, ToolSpec

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"
CSV_MIME = "text/csv"
TSV_MIME = "text/tab-separated-values"


def _to_rows(value: Any) -> list[list[Any]] | None:
    if isinstance(value, list):
        if not value:
            return []
        if all(isinstance(row, list) for row in value):
            return value
        return [value]
    if isinstance(value, str):
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return None
        return _to_rows(parsed)
    return None


def _json_safe_cell(value: Any) -> Any:
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, timedelta):
        return str(value)
    return value


def _json_safe_rows(rows: list[list[Any]]) -> list[list[Any]]:
    return [[_json_safe_cell(cell) for cell in row] for row in rows]


def _normalize_header(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return "column"
    raw = re.sub(r"[^a-z0-9]+", "_", raw)
    raw = re.sub(r"_+", "_", raw).strip("_")
    return raw or "column"


def _sheet_selector(value: Any) -> int | str:
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text:
        return 0
    if text.isdigit():
        return int(text)
    return text


def _qualify_range(sheet_name: str, range_string: str) -> str:
    return range_string if "!" in range_string else f"{sheet_name}!{range_string}"


class ExcelTool(BaseTool):
    name = "excel"

    def __init__(self, settings: Settings, recalc_engine: ExcelRecalcEngine | None = None) -> None:
        self.settings = settings
        self.recalc_engine = recalc_engine or ExcelRecalcEngine(timeout_seconds=settings.excel_recalc_timeout_seconds)

    def spec(self) -> ToolSpec:
        return ToolSpec(
            name=self.name,
            description="Professional Excel workflows for local spreadsheets in workspace.",
            input_schema={
                "type": "object",
                "properties": {
                    "action": {
                        "type": "string",
                        "enum": [
                            "create",
                            "list_sheets",
                            "read",
                            "write_cells",
                            "append_rows",
                            "add_sheet",
                            "convert",
                            "clean_table",
                            "set_number_format",
                            "set_style",
                            "add_comment",
                            "create_chart",
                            "recalc_validate",
                        ],
                    },
                    "path": {"type": "string"},
                    "output_path": {"type": "string"},
                    "sheet": {"type": "string"},
                    "sheet_name": {"type": "string"},
                    "range": {"type": "string"},
                    "cells": {"oneOf": [{"type": "object"}, {"type": "string"}]},
                    "rows": {"oneOf": [{"type": "array"}, {"type": "string"}]},
                    "normalize_headers": {"type": "boolean"},
                    "drop_empty_rows": {"type": "boolean"},
                    "drop_empty_cols": {"type": "boolean"},
                    "normalize_types": {"type": "boolean"},
                    "number_format": {"type": "string"},
                    "format": {"type": "string"},
                    "preset": {"type": "string", "enum": ["professional"]},
                    "font_name": {"type": "string"},
                    "font_size": {"type": "number"},
                    "bold": {"type": "boolean"},
                    "italic": {"type": "boolean"},
                    "font_color": {"type": "string"},
                    "fill_color": {"type": "string"},
                    "horizontal": {"type": "string"},
                    "vertical": {"type": "string"},
                    "cell": {"type": "string"},
                    "comment": {"type": "string"},
                    "author": {"type": "string"},
                    "chart_type": {"type": "string", "enum": ["line", "bar", "column"]},
                    "data_range": {"type": "string"},
                    "category_range": {"type": "string"},
                    "title": {"type": "string"},
                    "position": {"type": "string"},
                    "confirmed": {"type": "boolean"},
                },
                "required": ["action"],
            },
        )

    def _resolve_workspace_path(self, raw_path: str) -> Path:
        candidate = Path(raw_path).expanduser()
        if not candidate.is_absolute():
            candidate = self.settings.workspace / candidate
        resolved = candidate.resolve()
        workspace = self.settings.workspace.resolve()
        if workspace != resolved and workspace not in resolved.parents:
            raise PermissionError("path escapes workspace")
        return resolved

    @staticmethod
    def _artifact(path: Path) -> dict[str, Any]:
        suffix = path.suffix.lower()
        mime_type = XLSX_MIME
        if suffix == ".xlsm":
            mime_type = XLSM_MIME
        elif suffix == ".csv":
            mime_type = CSV_MIME
        elif suffix == ".tsv":
            mime_type = TSV_MIME
        return {
            "type": "document",
            "path": str(path),
            "file_name": path.name,
            "mime_type": mime_type,
        }

    @staticmethod
    def _write_preview(action: str, details: list[str]) -> str:
        lines = [
            "Excel write operation requires confirmation.",
            f"action={action}",
            *details,
            "Reply YES to proceed or NO to cancel.",
        ]
        return "\n".join(lines)

    @staticmethod
    def _load_cells(value: Any) -> dict[str, Any] | None:
        if isinstance(value, dict):
            return value
        if isinstance(value, str):
            try:
                parsed = json.loads(value)
            except json.JSONDecodeError:
                return None
            if isinstance(parsed, dict):
                return parsed
        return None

    @staticmethod
    def _iter_cells_from_range(worksheet, range_a1: str):  # noqa: ANN001
        cells = worksheet[range_a1]
        if isinstance(cells, tuple):
            if cells and isinstance(cells[0], tuple):
                for row in cells:
                    for cell in row:
                        yield cell
            else:
                for cell in cells:
                    yield cell
            return
        yield cells

    @staticmethod
    def _contains_formula(values: list[Any]) -> bool:
        for value in values:
            if isinstance(value, str) and value.strip().startswith("="):
                return True
        return False

    def _open_workbook(self, file_path: Path, *, data_only: bool = False):
        keep_vba = file_path.suffix.lower() == ".xlsm" and not data_only
        return load_workbook(file_path, data_only=data_only, keep_vba=keep_vba)

    def _recalc_summary_text(self, report: dict[str, Any]) -> str:
        total_errors = int(report.get("total_errors") or 0)
        formula_count = int(report.get("formula_count") or 0)
        summary = f"recalc_status={report.get('status','unknown')} formulas={formula_count} errors={total_errors}"
        error_summary = report.get("error_summary")
        if isinstance(error_summary, dict) and error_summary:
            lines = []
            for token, details in error_summary.items():
                if not isinstance(details, dict):
                    continue
                count = int(details.get("count") or 0)
                locations = details.get("locations")
                first = ""
                if isinstance(locations, list) and locations:
                    first = f" first={locations[0]}"
                lines.append(f"{token}:{count}{first}")
            if lines:
                summary += "\nerror_summary=" + ", ".join(lines)
        return summary

    def _auto_recalc_after_formula_write(self, file_path: Path) -> tuple[ToolResult | None, dict[str, Any] | None]:
        if not self.settings.excel_recalc_enabled:
            return None, None
        report = self.recalc_engine.recalc_and_validate(file_path)
        if not report.get("ok"):
            return ToolResult(ok=False, content=f"Formula validation failed: {report.get('error','unknown recalc error')}"), report

        total_errors = int(report.get("total_errors") or 0)
        if self.settings.excel_strict_formula_errors and total_errors > 0:
            return (
                ToolResult(
                    ok=False,
                    content=(
                        "Formula validation found Excel errors. Fix formulas and retry.\n"
                        f"{self._recalc_summary_text(report)}"
                    ),
                ),
                report,
            )
        return None, report

    async def run(self, args: dict[str, Any]) -> ToolResult:
        action = str(args.get("action") or "")
        raw_path = str(args.get("path") or "").strip()

        if action in {
            "create",
            "list_sheets",
            "read",
            "write_cells",
            "append_rows",
            "add_sheet",
            "convert",
            "clean_table",
            "set_number_format",
            "set_style",
            "add_comment",
            "create_chart",
            "recalc_validate",
        } and not raw_path:
            return ToolResult(ok=False, content="path is required")

        try:
            file_path = self._resolve_workspace_path(raw_path) if raw_path else None
        except PermissionError as exc:
            return ToolResult(ok=False, content=f"path rejected: {exc}")

        if action == "create":
            assert file_path is not None
            sheet_name = str(args.get("sheet_name") or "Sheet1").strip() or "Sheet1"
            if file_path.suffix.lower() not in {".xlsx", ".xlsm"}:
                file_path = file_path.with_suffix(".xlsx")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(
                        action,
                        [
                            f"path={file_path}",
                            f"sheet_name={sheet_name}",
                        ],
                    ),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )
            file_path.parent.mkdir(parents=True, exist_ok=True)
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            workbook.save(file_path)
            workbook.close()
            return ToolResult(
                ok=True,
                content=f"Workbook created.\npath={file_path}\nsheet={sheet_name}",
                artifacts=[self._artifact(file_path)],
            )

        if action == "list_sheets":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            workbook = self._open_workbook(file_path)
            try:
                names = workbook.sheetnames
            finally:
                workbook.close()
            return ToolResult(ok=True, content="Sheets:\n" + "\n".join(f"- {name}" for name in names))

        if action == "read":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            workbook = self._open_workbook(file_path, data_only=True)
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                range_a1 = str(args.get("range") or "").strip()
                if range_a1:
                    cells = ws[range_a1]
                    if not isinstance(cells, tuple):
                        rows = [[cells.value]]
                    else:
                        rows = [[cell.value for cell in row] for row in cells]
                    safe_rows = _json_safe_rows(rows)
                    return ToolResult(
                        ok=True,
                        content=(
                            f"Workbook read.\npath={file_path}\nsheet={sheet_name}\nrange={range_a1}\n"
                            f"values={json.dumps(safe_rows, ensure_ascii=False)}"
                        ),
                    )

                out_rows: list[list[Any]] = []
                for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 18)):
                    out_rows.append([cell.value for cell in row])
                safe_rows = _json_safe_rows(out_rows)
                return ToolResult(
                    ok=True,
                    content=(
                        f"Workbook read.\npath={file_path}\nsheet={sheet_name}\n"
                        f"values={json.dumps(safe_rows, ensure_ascii=False)}"
                    ),
                )
            finally:
                workbook.close()

        if action == "add_sheet":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            sheet_name = str(args.get("sheet_name") or "").strip()
            if not sheet_name:
                return ToolResult(ok=False, content="sheet_name is required")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"sheet_name={sheet_name}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )
            workbook = self._open_workbook(file_path)
            try:
                if sheet_name in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet already exists: {sheet_name}")
                workbook.create_sheet(sheet_name)
                workbook.save(file_path)
            finally:
                workbook.close()
            return ToolResult(
                ok=True,
                content=f"Sheet added.\npath={file_path}\nsheet={sheet_name}",
                artifacts=[self._artifact(file_path)],
            )

        if action == "write_cells":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            cells = self._load_cells(args.get("cells"))
            if not cells:
                return ToolResult(ok=False, content="cells must be a JSON object mapping cell => value")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"cells={len(cells)}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            workbook = self._open_workbook(file_path)
            formula_written = False
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                for cell_ref, value in cells.items():
                    ws[str(cell_ref)] = value
                    formula_written = formula_written or (
                        isinstance(value, str) and value.strip().startswith("=")
                    )
                workbook.save(file_path)
            finally:
                workbook.close()

            recalc_failure, recalc_report = (
                self._auto_recalc_after_formula_write(file_path) if formula_written else (None, None)
            )
            if recalc_failure is not None:
                return recalc_failure

            content = f"Cells updated.\npath={file_path}\nsheet={sheet_name}\nupdated={len(cells)}"
            if formula_written and recalc_report and recalc_report.get("ok"):
                content += "\n" + self._recalc_summary_text(recalc_report)

            return ToolResult(ok=True, content=content, artifacts=[self._artifact(file_path)])

        if action == "append_rows":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            rows = _to_rows(args.get("rows"))
            if rows is None:
                return ToolResult(ok=False, content="rows must be a 2D array (or JSON string)")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"rows={len(rows)}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            workbook = self._open_workbook(file_path)
            formula_written = False
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                for row in rows:
                    ws.append(row)
                    formula_written = formula_written or self._contains_formula(row)
                workbook.save(file_path)
            finally:
                workbook.close()

            recalc_failure, recalc_report = (
                self._auto_recalc_after_formula_write(file_path) if formula_written else (None, None)
            )
            if recalc_failure is not None:
                return recalc_failure

            content = f"Rows appended.\npath={file_path}\nsheet={sheet_name}\nrows={len(rows)}"
            if formula_written and recalc_report and recalc_report.get("ok"):
                content += "\n" + self._recalc_summary_text(recalc_report)
            return ToolResult(ok=True, content=content, artifacts=[self._artifact(file_path)])

        if action == "convert":
            assert file_path is not None
            if not file_path.exists() or not file_path.is_file():
                return ToolResult(ok=False, content=f"input file not found: {file_path}")
            raw_output = str(args.get("output_path") or "").strip()
            if raw_output:
                try:
                    output_path = self._resolve_workspace_path(raw_output)
                except PermissionError as exc:
                    return ToolResult(ok=False, content=f"output path rejected: {exc}")
            else:
                default_suffix = ".xlsx"
                if file_path.suffix.lower() in {".xlsx", ".xlsm"}:
                    default_suffix = ".csv"
                output_path = file_path.with_suffix(default_suffix)

            if output_path.suffix.lower() not in {".xlsx", ".xlsm", ".csv", ".tsv"}:
                return ToolResult(ok=False, content="output extension must be one of: .xlsx, .xlsm, .csv, .tsv")

            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"input={file_path}", f"output={output_path}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            input_suffix = file_path.suffix.lower()
            output_suffix = output_path.suffix.lower()
            sheet_name = str(args.get("sheet") or "").strip() or "Sheet1"
            output_path.parent.mkdir(parents=True, exist_ok=True)

            if input_suffix in {".xlsx", ".xlsm"} and output_suffix in {".csv", ".tsv"}:
                df = pd.read_excel(file_path, sheet_name=_sheet_selector(args.get("sheet")))
                sep = "\t" if output_suffix == ".tsv" else ","
                df.to_csv(output_path, index=False, sep=sep)
            elif input_suffix in {".csv", ".tsv"} and output_suffix in {".xlsx", ".xlsm"}:
                sep = "\t" if input_suffix == ".tsv" else ","
                df = pd.read_csv(file_path, sep=sep)
                df.to_excel(output_path, index=False, sheet_name=sheet_name)
            elif input_suffix in {".xlsx", ".xlsm"} and output_suffix in {".xlsx", ".xlsm"}:
                workbook = self._open_workbook(file_path)
                try:
                    workbook.save(output_path)
                finally:
                    workbook.close()
            else:
                return ToolResult(ok=False, content=f"unsupported conversion: {input_suffix} -> {output_suffix}")

            return ToolResult(
                ok=True,
                content=f"File converted.\ninput={file_path}\noutput={output_path}",
                artifacts=[self._artifact(output_path)],
            )

        if action == "clean_table":
            assert file_path is not None
            if not file_path.exists() or not file_path.is_file():
                return ToolResult(ok=False, content=f"input file not found: {file_path}")
            raw_output = str(args.get("output_path") or "").strip()
            try:
                output_path = self._resolve_workspace_path(raw_output) if raw_output else file_path
            except PermissionError as exc:
                return ToolResult(ok=False, content=f"output path rejected: {exc}")

            input_suffix = file_path.suffix.lower()
            output_suffix = output_path.suffix.lower()
            if input_suffix not in {".xlsx", ".xlsm", ".csv", ".tsv"}:
                return ToolResult(ok=False, content=f"unsupported input extension: {input_suffix}")
            if output_suffix not in {".xlsx", ".csv", ".tsv"}:
                return ToolResult(ok=False, content="clean_table output must be .xlsx, .csv, or .tsv")
            if input_suffix == ".xlsm" and output_path.resolve() == file_path.resolve():
                return ToolResult(
                    ok=False,
                    content="clean_table on .xlsm requires output_path to avoid macro-loss overwrite",
                )

            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"input={file_path}", f"output={output_path}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            sheet_arg = str(args.get("sheet") or "").strip()
            if input_suffix in {".xlsx", ".xlsm"}:
                df = pd.read_excel(file_path, sheet_name=sheet_arg or 0)
            else:
                df = pd.read_csv(file_path, sep="\t" if input_suffix == ".tsv" else ",")

            if bool(args.get("drop_empty_rows", True)):
                df = df.dropna(axis=0, how="all")
            if bool(args.get("drop_empty_cols", True)):
                df = df.dropna(axis=1, how="all")

            if bool(args.get("normalize_headers", True)):
                dedupe: dict[str, int] = {}
                headers: list[str] = []
                for col in df.columns:
                    base = _normalize_header(col)
                    idx = dedupe.get(base, 0)
                    dedupe[base] = idx + 1
                    headers.append(base if idx == 0 else f"{base}_{idx+1}")
                df.columns = headers

            if bool(args.get("normalize_types", False)):
                for col in df.columns:
                    series = df[col]
                    if series.dtype != "object":
                        continue
                    numeric = pd.to_numeric(series, errors="coerce")
                    if numeric.notna().sum() >= max(1, int(len(series) * 0.8)):
                        df[col] = numeric
                        continue
                    datelike = pd.to_datetime(series, errors="coerce")
                    if datelike.notna().sum() >= max(1, int(len(series) * 0.8)):
                        df[col] = datelike

            output_path.parent.mkdir(parents=True, exist_ok=True)
            if output_suffix == ".xlsx":
                df.to_excel(output_path, index=False, sheet_name=sheet_arg or "Cleaned")
            elif output_suffix == ".tsv":
                df.to_csv(output_path, index=False, sep="\t")
            else:
                df.to_csv(output_path, index=False)

            return ToolResult(
                ok=True,
                content=(
                    "Table cleaned.\n"
                    f"input={file_path}\noutput={output_path}\n"
                    f"rows={len(df.index)} cols={len(df.columns)}"
                ),
                artifacts=[self._artifact(output_path)],
            )

        if action == "set_number_format":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            range_a1 = str(args.get("range") or "").strip()
            number_format = str(args.get("number_format") or args.get("format") or "").strip()
            if not range_a1:
                return ToolResult(ok=False, content="range is required")
            if not number_format:
                return ToolResult(ok=False, content="number_format is required")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"range={range_a1}", f"format={number_format}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            workbook = self._open_workbook(file_path)
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                updated = 0
                for cell in self._iter_cells_from_range(ws, range_a1):
                    cell.number_format = number_format
                    updated += 1
                workbook.save(file_path)
            finally:
                workbook.close()

            return ToolResult(
                ok=True,
                content=f"Number format applied.\npath={file_path}\nsheet={sheet_name}\nrange={range_a1}\nupdated={updated}",
                artifacts=[self._artifact(file_path)],
            )

        if action == "set_style":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            range_a1 = str(args.get("range") or "").strip()
            if not range_a1:
                return ToolResult(ok=False, content="range is required")

            preset = str(args.get("preset") or "").strip().lower()
            font_name = str(args.get("font_name") or "").strip()
            font_size = args.get("font_size")
            bold = args.get("bold")
            italic = args.get("italic")
            font_color = str(args.get("font_color") or "").strip()
            fill_color = str(args.get("fill_color") or "").strip()
            horizontal = str(args.get("horizontal") or "").strip()
            vertical = str(args.get("vertical") or "").strip()

            if preset == "professional":
                if not font_name:
                    font_name = "Arial"
                if font_size is None:
                    font_size = 10

            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"range={range_a1}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            workbook = self._open_workbook(file_path)
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                updated = 0
                for cell in self._iter_cells_from_range(ws, range_a1):
                    style_font = copy(cell.font)
                    if font_name:
                        style_font.name = font_name
                    if isinstance(font_size, (int, float)):
                        style_font.sz = float(font_size)
                    if isinstance(bold, bool):
                        style_font.bold = bold
                    if isinstance(italic, bool):
                        style_font.italic = italic
                    if font_color:
                        style_font.color = font_color
                    cell.font = style_font

                    if fill_color:
                        cell.fill = PatternFill(fill_type="solid", start_color=fill_color, end_color=fill_color)

                    if horizontal or vertical:
                        align = copy(cell.alignment)
                        if horizontal:
                            align.horizontal = horizontal
                        if vertical:
                            align.vertical = vertical
                        cell.alignment = align

                    updated += 1
                workbook.save(file_path)
            finally:
                workbook.close()

            return ToolResult(
                ok=True,
                content=f"Style applied.\npath={file_path}\nsheet={sheet_name}\nrange={range_a1}\nupdated={updated}",
                artifacts=[self._artifact(file_path)],
            )

        if action == "add_comment":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            cell_ref = str(args.get("cell") or "").strip()
            comment_text = str(args.get("comment") or "").strip()
            author = str(args.get("author") or "Nexus").strip() or "Nexus"
            if not cell_ref:
                return ToolResult(ok=False, content="cell is required")
            if not comment_text:
                return ToolResult(ok=False, content="comment is required")

            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"cell={cell_ref}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            workbook = self._open_workbook(file_path)
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]
                ws[cell_ref].comment = Comment(comment_text, author)
                workbook.save(file_path)
            finally:
                workbook.close()

            return ToolResult(
                ok=True,
                content=f"Comment added.\npath={file_path}\nsheet={sheet_name}\ncell={cell_ref}",
                artifacts=[self._artifact(file_path)],
            )

        if action == "create_chart":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            data_range = str(args.get("data_range") or "").strip()
            chart_type = str(args.get("chart_type") or "").strip().lower() or "line"
            if not data_range:
                return ToolResult(ok=False, content="data_range is required")
            if chart_type not in {"line", "bar", "column"}:
                return ToolResult(ok=False, content=f"unsupported chart_type: {chart_type}")

            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}", f"data_range={data_range}", f"chart_type={chart_type}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )

            category_range = str(args.get("category_range") or "").strip()
            title = str(args.get("title") or "").strip() or "Chart"
            position = str(args.get("position") or "E2").strip() or "E2"

            workbook = self._open_workbook(file_path)
            try:
                sheet_name = str(args.get("sheet") or "").strip() or workbook.sheetnames[0]
                if sheet_name not in workbook.sheetnames:
                    return ToolResult(ok=False, content=f"sheet not found: {sheet_name}")
                ws = workbook[sheet_name]

                if chart_type == "line":
                    chart = LineChart()
                else:
                    chart = BarChart()
                    if chart_type == "column":
                        chart.type = "col"
                    else:
                        chart.type = "bar"

                chart.title = title
                data_ref = Reference(ws, range_string=_qualify_range(sheet_name, data_range))
                chart.add_data(data_ref, titles_from_data=True)

                if category_range:
                    cat_ref = Reference(ws, range_string=_qualify_range(sheet_name, category_range))
                    chart.set_categories(cat_ref)

                ws.add_chart(chart, position)
                workbook.save(file_path)
            finally:
                workbook.close()

            return ToolResult(
                ok=True,
                content=(
                    "Chart created.\n"
                    f"path={file_path}\nsheet={sheet_name}\nchart_type={chart_type}\n"
                    f"data_range={data_range}\nposition={position}"
                ),
                artifacts=[self._artifact(file_path)],
            )

        if action == "recalc_validate":
            assert file_path is not None
            if not file_path.exists():
                return ToolResult(ok=False, content=f"workbook not found: {file_path}")
            if not args.get("confirmed"):
                return ToolResult(
                    ok=False,
                    content=self._write_preview(action, [f"path={file_path}"]),
                    requires_confirmation=True,
                    risk_level="high",
                    proposed_action={"action": action, **args},
                )
            report = self.recalc_engine.recalc_and_validate(file_path)
            if not report.get("ok"):
                return ToolResult(ok=False, content=f"Recalc failed: {report.get('error','unknown error')}")
            total_errors = int(report.get("total_errors") or 0)
            ok = (not self.settings.excel_strict_formula_errors) or total_errors == 0
            return ToolResult(
                ok=ok,
                content=(
                    "Recalculation complete.\n"
                    f"path={file_path}\n{self._recalc_summary_text(report)}"
                ),
                artifacts=[self._artifact(file_path)] if ok else [],
            )

        return ToolResult(ok=False, content=f"Unsupported action: {action}")
